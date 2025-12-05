"""
Advanced Reporting Service for PDF and Excel report generation
"""
from typing import Dict, List, Optional
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
import logging
import io
import json

from app.models.site import SharePointSite
from app.models.audit import AuditLog
from app.models.access_review import AccessReviewCycle

logger = logging.getLogger(__name__)


class ReportingService:
    """Service for generating advanced reports"""
    
    def __init__(self, db: Session):
        self.db = db
    
    async def generate_executive_summary(
        self,
        start_date: datetime,
        end_date: datetime,
        format: str = "pdf"
    ) -> bytes:
        """
        Generate executive summary report
        
        Args:
            start_date: Report start date
            end_date: Report end date
            format: 'pdf' or 'excel'
        
        Returns:
            Report file bytes
        """
        logger.info(f"Generating executive summary from {start_date} to {end_date}")
        
        # Gather statistics
        total_sites = self.db.query(SharePointSite).filter(
            SharePointSite.is_archived == False
        ).count()
        
        audit_events = self.db.query(AuditLog).filter(
            AuditLog.event_datetime >= start_date,
            AuditLog.event_datetime <= end_date
        ).count()
        
        completed_reviews = self.db.query(AccessReviewCycle).filter(
            AccessReviewCycle.certified_date >= start_date,
            AccessReviewCycle.certified_date <= end_date
        ).count()
        
        data = {
            "report_title": "Governance Executive Summary",
            "generated_at": datetime.utcnow().isoformat(),
            "period": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat()
            },
            "metrics": {
                "total_active_sites": total_sites,
                "audit_events": audit_events,
                "completed_reviews": completed_reviews,
            }
        }
        
        if format == "pdf":
            return self._generate_pdf_report(data)
        else:
            return self._generate_excel_report(data)
    
    def _generate_pdf_report(self, data: Dict) -> bytes:
        """
        Generate PDF report using reportlab
        
        Note: This is a placeholder. Actual implementation would use reportlab
        """
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import inch
            
            buffer = io.BytesIO()
            pdf = canvas.Canvas(buffer, pagesize=letter)
            
            # Title
            pdf.setFont("Helvetica-Bold", 24)
            pdf.drawString(inch, 10*inch, data["report_title"])
            
            # Metrics
            pdf.setFont("Helvetica", 12)
            y = 9*inch
            pdf.drawString(inch, y, f"Generated: {data['generated_at']}")
            
            y -= 0.5*inch
            pdf.drawString(inch, y, f"Period: {data['period']['start']} to {data['period']['end']}")
            
            y -= inch
            pdf.setFont("Helvetica-Bold", 14)
            pdf.drawString(inch, y, "Summary Metrics:")
            
            y -= 0.5*inch
            pdf.setFont("Helvetica", 12)
            for key, value in data["metrics"].items():
                pdf.drawString(inch * 1.5, y, f"{key.replace('_', ' ').title()}: {value}")
                y -= 0.3*inch
            
            pdf.save()
            return buffer.getvalue()
            
        except ImportError:
            logger.warning("reportlab not available, returning JSON")
            return json.dumps(data, indent=2).encode()
    
    def _generate_excel_report(self, data: Dict) -> bytes:
        """
        Generate Excel report using openpyxl
        
        Note: This is a placeholder. Actual implementation would use openpyxl
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Executive Summary"
            
            # Write data
            ws['A1'] = data["report_title"]
            ws['A2'] = f"Generated: {data['generated_at']}"
            ws['A3'] = f"Period: {data['period']['start']} to {data['period']['end']}"
            
            row = 5
            ws[f'A{row}'] = "Metric"
            ws[f'B{row}'] = "Value"
            
            row += 1
            for key, value in data["metrics"].items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = value
                row += 1
            
            buffer = io.BytesIO()
            wb.save(buffer)
            return buffer.getvalue()
            
        except ImportError:
            logger.warning("openpyxl not available, returning JSON")
            return json.dumps(data, indent=2).encode()
    
    async def schedule_report(
        self,
        report_type: str,
        schedule_cron: str,
        recipients: List[str],
        format: str = "pdf"
    ) -> Dict:
        """
        Schedule a recurring report
        
        Args:
            report_type: Type of report
            schedule_cron: Cron expression
            recipients: List of email addresses
            format: Report format
        
        Returns:
            Schedule confirmation
        """
        logger.info(f"Scheduling {report_type} report with cron: {schedule_cron}")
        
        # TODO: Add to APScheduler with cron trigger
        # This would require storing scheduled reports in database
        
        return {
            "report_type": report_type,
            "schedule": schedule_cron,
            "recipients": recipients,
            "format": format,
            "status": "scheduled"
        }


def get_reporting_service(db: Session) -> ReportingService:
    """Dependency to get reporting service"""
    return ReportingService(db)
